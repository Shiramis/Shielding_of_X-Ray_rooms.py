from tkinter import *
from tkinter import ttk
import pandas as pd
import math
import os
from openpyxl import load_workbook

# Get the current script directory
script_dir = os.path.dirname(__file__)

excel_path = os.path.join(script_dir, '..', 'Data Shielding.xlsx')

wb = load_workbook(excel_path)

class departprimsec():

    def choosetype(self,e,nr, t):
        if self.d["vselroom "+str(t)].get() == "CT Room":
            self.depCTcal(e, nr, nb, t)
        else:
            Ktotal = 0
            ce = None
            for i in range(0,self.d[f"num_barriers_var {e}{nr}"].get()):
                K = self.calkerma(e, nr, i, t)
                if K is not None:
                    Ktotal += K
                print ("Kt="+str(Ktotal))
                if f"setv {e}{nr}{i}" in self.d:
                    if self.d[f"setv {e}{nr}{i}"].get() == 1:
                        ce = e
                        cnr = nr
                        ci = i
            if self.d["dikeent " + str(e) + nr].get() == "":
                self.need.append("Design goal Kerma")
                self.depcalc(e, nr, i, t)
            else:
                if ce is None:
                    self.need.append("You must set a Workload Distribution")
                    self.depcalc(e, nr, i, t)
                else:
                    self.B = (float(self.d["dikeent " + str(e) + nr].get()))/Ktotal
            if ce is None:
                self.need.append("You must set a Workload Distribution")
                self.depcalc(e, nr, i, t)
            else:
                self.depcalc(ce, cnr, ci, t)

    def calkerma(self, e, nr, i, t):
        if self.d[f"entryd {e}{nr}{i}"].get() != "":
            D = float(self.d[f"entryd {e}{nr}{i}"].get()) # Distance (entryd)
            if self.barn[f"lab_bar {e}{nr}"].cget("text") == "Floor": #from fig. 4.4 NCRP 147
                D += 1.7
            elif self.barn[f"lab_bar {e}{nr}"].cget("text") == "Ceiling":
                D += 0.5
            else:
                D += 0.3
        else:
            D = None
        # ======== Workload ===========
        if f"othwork {e}{nr}{i}" in self.d:
            if self.d[f"oworkvar {e}{nr}{i}"] == 1: #If checks different workload
                if self.d[f"workv {e}{nr}{i}"].get() == 2:
                    if self.d[f"numpapwe {e}{nr}{i}"].get() != '':
                        n = int(self.d[f"numpapwe {e}{nr}{i}"].get())  # Number of patients (numpapwe)
                    else:
                        n =None
                elif self.d[f"workv {e}{nr}{i}"].get() == 1:# Workload option
                    ws = wb['Workload']
                    for x in range(1, 13):
                        if self.d[f"vsexroom {e}{nr}{i}"].get() == ws['A' + str(x)].value:  # X-ray room selection (vsexroom)
                            if self.d[f"worentry {e}{nr}{i}"].get() != '':
                                n = float(self.d[f"worentry {e}{nr}{i}"].get()) / ws['B' + str(x)].value  # Workload entry (worentry)
                            else:
                                n = None
                    if self.d[f"vselxray {e}{nr}{i}"].get() == 2:  # X-ray selection (vselxray)
                        if self.d[f"worentry {e}{nr}{i}"].get() != '':
                            n = float(self.d[f"worentry {e}{nr}{i}"].get()) / 2.5  # Workload entry (worentry)
                        else:
                            n = None
            else:
                # The general workload
                if self.d[f"vrawork {t}"].get() == 2:
                    if self.d[f"numpapwe {t}"].get() != '':
                        n = int(self.d[f"numpapwe {t}"].get())
                    else:
                        n = None
                elif self.d[f"vrawork {t}"].get() == 1:
                    ws = wb['Workload']
                    for x in range(1, 13):
                        if self.d[f"vsexroom {e}{nr}{i}"].get() == ws['A' + str(x)].value:
                            if self.d[f"worentry {t}"].get() != '':
                                n = float(self.d[f"worentry {t}"].get()) / ws['B' + str(x)].value
                            else:
                                n = None
                    if self.d[f"vselxray {t}"].get() == 2:
                        if self.d[f"worentry {t}"].get() != '':
                            n = float(self.d[f"worentry {t}"].get()) / 2.5
                        else:
                            n = None
                else:
                    n = None
            
        else:
            # The general workload
            if self.d[f"vrawork {t}"].get() == 2:
                if self.d[f"numpapwe {t}"].get() != '':
                    n = int(self.d[f"numpapwe {t}"].get())
                else:
                    n = None
            elif self.d[f"vrawork {t}"].get() == 1:
                ws = wb['Workload']
                for x in range(1, 13):
                    if self.d[f"vsexroom {e}{nr}{i}"].get() == ws['A' + str(x)].value:
                        if self.d[f"worentry {t}"].get() != '':
                            n = float(self.d[f"worentry {t}"].get()) / ws['B' + str(x)].value
                        else:
                            n = None
                if self.d[f"vselxray {t}"].get() == 2:
                    if self.d[f"worentry {t}"].get() != '':
                        n = float(self.d[f"worentry {t}"].get()) / 2.5
                    else:
                        n = None
            else:
                n = None
        # ======== Occupancy Factor T ===========
        if self.d[f"vraoccup {e}{nr}"].get() == 1:  # Occupancy option (vraoccup)
            if self.d[f"occupentry {e}{nr}"].get() != '':
                T = float(self.d[f"occupentry {e}{nr}"].get())  # Occupancy entry (occupentry)
            else:
                T = None
        elif self.d[f"vraoccup {e}{nr}"].get() == 2:
            ws = wb['Occupancy Factor ( T )']
            for x in range(2, 32):
                if self.d[f"vselocation {e}{nr}"].get() == ws['A' + str(x)].value:  # Location selection (vselocation)
                    T = float(ws['B' + str(x)].value)
        else:
            T = None
        # ======== K1 "air kerma" ===========
        if self.d[f"radiob_w {e}{nr}{i}"].get() == 1:
        # primary unshielded air kerma
            if self.d[f"entk {e}{nr}{i}"].get() != "":
                K1 = float(self.d[f"entk {e}{nr}{i}"].get())  # Entered kerma value (entk)
            else:
                K1 = None
        elif self.d[f"radiob_w {e}{nr}{i}"].get() == 2:
            if self.d[f"unairkerv {e}{nr}{i}"].get() == 1: #Selects fron NCRP 147
                # Secondary air kerma calculations based on the 'airkerv' value
                ws = wb["Uns Air Kerma"]
                room_type = self.d[f"vsexroom {e}{nr}{i}"].get()  # Fetching X-ray room type
                if self.d[f"airkerv {e}{nr}{i}"].get() == 1:
                    # Leakage air kerma calculation
                    if self.d[f"radiob_leak {e}{nr}{i}"].get() == 1:
                        for j in range(1, 11):
                            if room_type == ws['A' + str(j)].value:
                                K1 = float(ws['G' + str(j)].value)  # Leak & side scatter
                    elif self.d[f"radiob_leak {e}{nr}{i}"].get() == 2:
                        for j in range(1, 11):
                            if room_type == ws['A' + str(j)].value:
                                K1 = float(ws['I' + str(j)].value)  # Leakage and Forward/ Backscatter (Ksec)
                    else:
                        for j in range(1, 11):
                            if room_type == ws['A' + str(j)].value:
                                K1 = float(ws['E' + str(j)].value)  # Only Leakage

                elif self.d[f"airkerv {e}{nr}{i}"].get() == 2:
                    # Side-Scatter
                    for j in range(1, 11):
                        if room_type == ws['A' + str(j)].value:
                            K1 = float(ws['F' + str(j)].value)
                elif self.d[f"airkerv {e}{nr}{i}"].get() == 3:
                    # Forward/ Backscatter
                    for j in range(1, 11):
                        if room_type == ws['A' + str(j)].value:
                            K1 = float(ws['H' + str(j)].value)
            elif self.d[f"unairkerv {e}{nr}{i}"].get() == 2:
                # User-entered air kerma
                if self.d[f"entk {e}{nr}{i}"].get() != "":
                    K1 = float(self.d[f"entk {e}{nr}{i}"].get())  # Entered kerma value (entk)
                else:
                    K1 = None
            else:
                K1 = None
        else:
            K1 = None
        # ======== Use Factor ===========
        if self.d[f"radiob_w {e}{nr}{i}"].get() == 1:
            if self.d[f"use_ent {e}{nr}{i}"].get()!= "":
                Us = float(self.d[f"use_ent {e}{nr}{i}"].get())  # Use factor (use_ent)
            else:
                Us =None
        elif self.d[f"radiob_w {e}{nr}{i}"].get() == 2:
            Us = 1
        else:
            Us = None
        #====Missed variables========
        self.need = []
        if n is None:
            self.need.append("Workload")
            print('w')
        if Us is None:
            self.need.append("Use Factor (U)")
            print('us')
        if T is None:
            self.need.append("Occupancy Factor (T)")
            print('T')
        if K1 is None:
            self.need.append("Unshielded Air Kerma")
            print('k')
        if D is None:
            self.need.append("Distance from the Source to the Barrier (m)")
            print("d")
        if n and Us and T and K1 and D is not None:
            # ========== Calculate Kerma ==========
            K = n * Us * T * K1 / (D ** 2)
            return K

        return None  # Return None if there are missing variables

    def depcalc(self, e, nr, i, t):  # Barrier calculations

        title_key = f"titleresul {e}{nr}"
        # Destroy existing material labels if they exist
        if self.d.get(title_key) is not None:
            for o in range(1, 7):
                mat_key = f"resmat {o}{e}"
                if self.res.get(mat_key) is not None:
                    self.res[mat_key].destroy()
        else:
            for o in range(1, 7):
                self.res[f"resmat {o}{e}"] = None

        # Iterate over materials and perform calculations
        for o in range(1, self.d[f"vnumbmat {e}{nr}"].get() + 1):
            material = self.d[f"vmater {e}{o}{nr}"].get()
            if self.need:
                if material == "Select Material":
                    self.need.append("Select Material")
                self.need = list(dict.fromkeys(self.need))
                self.thm[f"xbar {e}{o}{nr}"] = "\n".join(self.need)
                self.display_results(e, o, nr, t)
            else:
                if self.d[f"radiob_w {e}{nr}{i}"].get() == 1:
                    ws = wb['prim abc']
                    for x in range(2, 39):
                        if self.d[f"vsexroom {e}{nr}{i}"].get() == ws[f'A{x}'].value:
                            self.al = ws[f'B{x}'].value
                            self.bl = ws[f'C{x}'].value
                            self.cl = float(ws[f'D{x}'].value)
                            self.ac = ws[f'E{x}'].value
                            self.bc = ws[f'F{x}'].value
                            self.cc = float(ws[f'G{x}'].value)
                            self.ag = ws[f'H{x}'].value
                            self.bg = ws[f'I{x}'].value
                            self.cg = float(ws[f'J{x}'].value)
                            self.ast = ws[f'K{x}'].value
                            self.bs = ws[f'L{x}'].value
                            self.cs = float(ws[f'M{x}'].value)
                            self.apg = ws[f'N{x}'].value
                            self.bpg = ws[f'O{x}'].value
                            self.cpg = float(ws[f'P{x}'].value)
                            self.aw = ws[f'Q{x}'].value
                            self.bw = ws[f'R{x}'].value
                            self.cw = float(ws[f'S{x}'].value)

                    # Preshielding calculations
                    if self.d[f"preshvar {e}{nr}{i}"].get() == 1:
                        ws = wb["Equiv. thickness of prim pres"]
                        if self.d[f"radiob_pre {e}{nr}{i}"].get() == 1:
                            self.xlead = float(ws['B3'].value)
                            self.xconc = float(ws['C3'].value)
                            self.xsteel = float(ws['D3'].value)
                        elif self.d[f"radiob_pre {e}{nr}{i}"].get() == 2:
                            self.xlead = float(ws['B4'].value)
                            self.xconc = float(ws['C4'].value)
                            self.xsteel = float(ws['D4'].value)
                    else:
                        self.xlead, self.xconc, self.xsteel = 0, 0, 0

                elif self.d[f"radiob_w {e}{nr}{i}"].get() == 2:
                    self.xlead, self.xconc, self.xsteel = 0, 0, 0
                    ws = wb['sec abc']
                    for x in range(3, 18):
                        if self.d[f"vsexroom {e}{nr}{i}"].get() == ws[f'A{x}'].value:
                            self.al = ws[f'B{x}'].value
                            self.bl = ws[f'C{x}'].value
                            self.cl = float(ws[f'D{x}'].value)
                            self.ac = ws[f'E{x}'].value
                            self.bc = ws[f'F{x}'].value
                            self.cc = float(ws[f'G{x}'].value)
                            self.ag = ws[f'H{x}'].value
                            self.bg = ws[f'I{x}'].value
                            self.cg = float(ws[f'J{x}'].value)
                            self.ast = ws[f'K{x}'].value
                            self.bs = ws[f'L{x}'].value
                            self.cs = float(ws[f'M{x}'].value)
                            self.apg = ws[f'N{x}'].value
                            self.bpg = ws[f'O{x}'].value
                            self.cpg = float(ws[f'P{x}'].value)
                            self.aw = ws[f'Q{x}'].value
                            self.bw = ws[f'R{x}'].value
                            self.cw = float(ws[f'S{x}'].value)

                # Material thickness calculations
                if material == "Select Material":
                    self.thm[f"xbar {e}{o}{nr}"] = "Select Material"
                elif material == "Lead":
                    self.thm[f"xbar {e}{o}{nr}"] = (1 / (self.al * self.cl)) * math.log(
                        (self.B ** (-self.cl) + self.bl / self.al) / (1 + self.bl / self.al)) - self.xlead
                elif material == "Concrete":
                    self.thm[f"xbar {e}{o}{nr}"] = (1 / (self.ac * self.cc)) * math.log(
                        (self.B ** (-self.cc) + self.bc / self.ac) / (1 + self.bc / self.ac)) - self.xconc
                elif material == "Gypsum Wallboard":
                    self.thm[f"xbar {e}{o}{nr}"] = (1 / (self.ag * self.cg)) * math.log(
                        (self.B ** (-self.cg) + self.bg / self.ag) / (1 + self.bg / self.ag))
                elif material == "Steel":
                    self.thm[f"xbar {e}{o}{nr}"] = (1 / (self.ast * self.cs)) * math.log(
                        (self.B ** (-self.cs) + self.bs / self.ast) / (1 + self.bs / self.ast)) - self.xsteel
                elif material == "Plate Glass":
                    self.thm[f"xbar {e}{o}{nr}"] = (1 / (self.apg * self.cpg)) * math.log(
                        (self.B ** (-self.cpg) + self.bpg / self.apg) / (1 + self.bpg / self.apg))
                elif material == "Wood":
                    self.thm[f"xbar {e}{o}{nr}"] = (1 / (self.aw * self.cw)) * math.log(
                        (self.B ** (-self.cw) + self.bw / self.aw) / (1 + self.bw / self.aw))

                self.display_results(e, o, nr, t)

        # If there's no existing title result, create a new title
        if self.d.get(title_key) is None:
            self.op = 0
            self.d[title_key] = ttk.Label(self.d[f"resultframe {t}{nr}"], style="AL.TLabel",
                text=self.barn[f"lab_bar {e}{nr}"].cget("text") + ": ")
            self.d[title_key].grid(row=str(self.ep - self.op), column=0, pady=3, padx=3, sticky="s")
            self.d[f"spot {e}{nr}"] = self.ep
        else:
            # If it exists, update it
            self.d[title_key].destroy()
            self.d[title_key] = ttk.Label(self.d[f"resultframe {t}{nr}"], style="AL.TLabel",
                text=self.barn[f"lab_bar {e}{nr}"].cget("text") + ": ")
            self.d[title_key].grid(row=str(self.d[f"spot {e}{nr}"]), column=0, pady=3, padx=3, sticky="s")

        # Store the results for further use
        if isinstance(self.thm[f'xbar {e}{o}{nr}'], float):
            self.rdata = {self.barn[f"lab_bar {e}{nr}"].cget("text"): {
                self.d[f"vmater {e}{o}{nr}"].get(): str(round(self.thm[f"xbar {e}{o}{nr}"], 3))}}

        self.ep += 1

    def depCTcal(self, e, nr, nb, t):
        if self.d["titleresul " + str(e)+nr] is not None:
            for o in range (1,7):
                if self.res["resmat " + str(o)+str(e)] is not None:
                    self.res["resmat " + str(o)+str(e)].destroy()
        else:
            for o in range(1, 7):
                self.res["resmat {0}".format(str(o)) + str(e)] = None
        for o in range(1, self.d["vnumbmat " + str(e) + nr].get() + 1):
            k1sec_body = float(1.2 * (3 * 10 ** -4) * (1.4 * float(self.d["dlpb_var "+ str(t)].get())))
            k1sec_head = float((9 * 10 ** -5) * (1.4 * float(self.d['dlph_var '+ str(t)].get())))
            self.d["K "+self.barn["lab_bar " + str(e) + nr].cget("text")+nr] = float((1 / float(self.d["dist_var "+ str(e) + nr].get())) ** 2
                                                                     * ((self.d["bp_var "+str(t)].get() * k1sec_body)
                                                                        + (self.d["hp_var " + str(t)].get() * k1sec_head)))
            B = float(self.d["sh_var " + str(e) + nr].get() / float(self.d["K "+self.barn["lab_bar " + str(e) + nr].cget("text")+nr]))
            print(B)
            if self.d["vmater "+str(e)+str(o)+nr].get() == "Lead":
                if self.d["kvp_var "+ str(t)].get() == 120:
                    a = 2.246
                    b = 5.73
                    c = 0.547
                else:
                    a = 2.009
                    b = 3.99
                    c = 0.342
            elif self.d["vmater "+str(e)+str(o)+nr].get() == "Concrete":
                if self.d["kvp_var "+ str(t)].get() == 120:
                    a = 0.0383
                    b = 0.0142
                    c = 0.658
                else:
                    a = 0.0336
                    b = 0.0122
                    c = 0.519
            self.thm["xbar "+str(e)+str(o)+nr] = float((1 / (a * c)) * math.log((B ** -c + (b / a)) / (1 + (b / a))))

            if self.d["titleresul " + str(e) + nr] is None:
                self.res["resmat " + str(o) + str(e)] = ttk.Label(self.d["resultframe " + str(t) + nr],
                                                                  style="AL.TLabel", text=self.d[
                                                                                              "vmater " + str(e) + str(
                                                                                                  o) + nr].get() + ": " + str(
                        round(self.thm["xbar " + str(e) + str(o) + nr], 2)) + " mm")
                if o < 3:
                    self.res["resmat " + str(o) + str(e)].grid(row=str(self.ep), column=o, pady=3, padx=3, sticky="w")
                    self.op= 0
                elif 2 < o < 5:
                    if o == 3:
                        self.ep += 1
                        self.op+= 1
                    self.res["resmat " + str(o) + str(e)].grid(row=str(self.ep), column=o - 2, pady=3, padx=3,
                                                               sticky="w")
                else:
                    if o == 5:
                        self.ep += 1
                        self.op+= 1
                    self.res["resmat " + str(o) + str(e)].grid(row=str(self.ep), column=o - 4, pady=3, padx=3,
                                                               sticky="w")
            else:
                self.res["resmat " + str(o) + str(e)] = ttk.Label(self.d["resultframe " + str(t) + nr],
                                                                  style="AL.TLabel", text=self.d[
                                                                                              "vmater " + str(e) + str(
                                                                                                  o) + nr].get() + ": " + str(
                        round(self.thm["xbar " + str(e) + str(o) + nr], 2)) + " mm")
                if o < 3:
                    self.res["resmat " + str(o) + str(e)].grid(row=str(self.d["spot " + str(e)]), column=o, pady=3,
                                                               padx=3, sticky="w")
                elif 2 < o < 5:
                    self.res["resmat " + str(o) + str(e)].grid(row=str(self.d["spot " + str(e)] + 1), column=o - 2,
                                                               pady=3, padx=3, sticky="w")
                else:
                    self.res["resmat " + str(o) + str(e)].grid(row=str(self.d["spot " + str(e)] + 2), column=o - 4,
                                                               pady=3, padx=3, sticky="w")

        if self.d["titleresul " + str(e) + nr] is None:
            self.d["titleresul " + str(e) + nr] = ttk.Label(self.d["resultframe " + str(t) + nr], style="AL.TLabel",
                                                            text=self.barn["lab_bar " + str(e) + nr].cget(
                                                                "text") + ": ")
            self.d["titleresul " + str(e) + nr].grid(row=str(self.ep - op), column=0, pady=3, padx=3, sticky="s")
            self.d["spot {0}".format(str(e))] = self.ep
        else:
            self.d["titleresul " + str(e) + nr].destroy()
            self.d["titleresul " + str(e) + nr] = ttk.Label(self.d["resultframe " + str(t) + nr], style="AL.TLabel",
                                                            text=self.barn["lab_bar " + str(e) + nr].cget(
                                                                "text") + ": ")
            self.d["titleresul " + str(e) + nr].grid(row=str(self.d["spot " + str(e)]), column=0, pady=3, padx=3,
                                                     sticky="s")

        print(self.d["K " + self.barn["lab_bar " + str(e) + nr].cget("text") + nr])
        self.ep += 1

    def display_results(self, e, o, nr, t):
        title_key = f"titleresul {e}{nr}"
        spot_key = f"spot {e}{nr}"
        if spot_key not in self.d:
            self.d[spot_key] = self.ep  # or another default value
        if self.d[title_key] is None:
            if isinstance(self.thm[f'xbar {e}{o}{nr}'], str):
                if self.need:
                    self.res[f"resmat {o}{e}"] = Text(self.d[f"resultframe {t}{nr}"], height=9, width=40,
                                                      borderwidth=0, background="#f7faf9", font='Helvetica 12')
                    # Insert "You must enter:" in black
                    self.res[f"resmat {o}{e}"].insert(END, "You must enter:", "black")
                    # Insert the rest of the text in red
                    self.res[f"resmat {o}{e}"].insert(END, f"\n{self.thm[f'xbar {e}{o}{nr}']}", "red")
                    # Configure the tags for text styles
                    self.res[f"resmat {o}{e}"].tag_configure("black", foreground="black")
                    self.res[f"resmat {o}{e}"].tag_configure("red", foreground='#f71616')
                    self.res[f"resmat {o}{e}"].config(state=DISABLED)
                else:
                    self.res[f"resmat {o}{e}"] = ttk.Label(self.d[f"resultframe {t}{nr}"], style="RL.TLabel",
                                                           text=f"{self.thm[f'xbar {e}{o}{nr}']}")
            else:
                self.res[f"resmat {o}{e}"] = ttk.Label(self.d[f"resultframe {t}{nr}"], style="AL.TLabel",
                                                       text=f"{self.d[f'vmater {e}{o}{nr}'].get()}: {round(self.thm[f'xbar {e}{o}{nr}'], 2)} mm")

            # Handle grid placement and manage row/column positions with op
            if o < 3:
                self.res[f"resmat {o}{e}"].grid(row=self.ep, column=o, pady=3, padx=3, sticky="w")
                self.op = 0  # Reset op if material is within the first 2 columns
            elif 2 < o < 5:
                if o == 3:
                    self.ep += 1  # Move to the next row for column 3
                    self.op += 1
                self.res[f"resmat {o}{e}"].grid(row=self.ep, column=o - 2, pady=3, padx=3, sticky="w")
            else:
                if o == 5:
                    self.ep += 1  # Move to a new row for column 5 and beyond
                    self.op += 1
                self.res[f"resmat {o}{e}"].grid(row=self.ep, column=o - 4, pady=3, padx=3, sticky="w")
        else:
            # If the title already exists, update it
            if isinstance(self.thm[f'xbar {e}{o}{nr}'], str):
                if self.need:
                    self.res[f"resmat {o}{e}"] = Text(self.d[f"resultframe {t}{nr}"], height=9, width=40,
                                                      borderwidth=0, background="#f7faf9", font='Helvetica 12')
                    # Insert "You must enter:" in black
                    self.res[f"resmat {o}{e}"].insert(END, "You must enter:", "black")
                    # Insert the rest of the text in red
                    self.res[f"resmat {o}{e}"].insert(END, f"\n{self.thm[f'xbar {e}{o}{nr}']}", "red")
                    # Configure the tags for text styles
                    self.res[f"resmat {o}{e}"].tag_configure("black", foreground="black")
                    self.res[f"resmat {o}{e}"].tag_configure("red", foreground='#f71616')
                    self.res[f"resmat {o}{e}"].config(state=DISABLED)
                else:
                    self.res[f"resmat {o}{e}"] = ttk.Label(self.d[f"resultframe {t}{nr}"], style="RL.TLabel",
                                                           text=f"{self.thm[f'xbar {e}{o}{nr}']}")
            else:
                self.res[f"resmat {o}{e}"] = ttk.Label(self.d[f"resultframe {t}{nr}"], style="AL.TLabel",
                                                       text=f"{self.d[f'vmater {e}{o}{nr}'].get()}: {round(self.thm[f'xbar {e}{o}{nr}'], 2)} mm")

            # Use stored row information (spot) for placing result labels
            if o < 3:
                self.res[f"resmat {o}{e}"].grid(row=self.d[spot_key], column=o, pady=3, padx=3, sticky="w")
            elif 2 < o < 5:
                self.res[f"resmat {o}{e}"].grid(row=self.d[spot_key] + 1, column=o - 2, pady=3, padx=3, sticky="w")
            else:
                self.res[f"resmat {o}{e}"].grid(row=self.d[spot_key] + 2, column=o - 4, pady=3, padx=3, sticky="w")

