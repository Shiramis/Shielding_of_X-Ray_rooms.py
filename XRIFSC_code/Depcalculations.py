from tkinter import *
from tkinter import ttk
import pandas as pd
import math
import os
from numpy.ma.core import append
from openpyxl import load_workbook
import os
import sys


def resource_path(relative_path):
    """ Get the absolute path to the resource, works for dev and for PyInstaller """
    try:
        # PyInstaller creates a temporary folder and stores path in _MEIPASS
        base_path = sys._MEIPASS
    except AttributeError:
        base_path = os.path.abspath("..")

    return os.path.join(base_path, relative_path)

# Update the excel file path
excel_path = resource_path('Data Shielding.xlsx')
if not os.path.exists(excel_path):
    print(f"Excel file not found at: {excel_path}")

wb = load_workbook(excel_path)

class departprimsec():

    def choosetype(self,e,nr, t):
        if self.var["vselroom "+str(t)].get() == "CT Room":
            self.depCTcal(e, nr, t)
        else:
            Ktotal = 0
            ce = None
            for i in range(0,self.var[f"num_barriers_var {e}{nr}"].get()):
                K = self.calkerma(e, nr, i, t)
                if K is not None:
                    Ktotal += K
                print ("Kt="+str(Ktotal))
                if f"setv {e}{nr}{i}" in self.var:
                    if self.var[f"setv {e}{nr}{i}"].get() == 1:
                        ce = e
                        cnr = nr
                        ci = i
            if self.d["dikeent " + str(e) + nr].get() == "":
                self.need.append("Design goal Kerma")
                self.depcalc(e, nr, i, t)
            else:
                if ce is None:
                    self.need.append("You must set a Workload Distribution")
                    self.depcalc(e, nr, i, t)

                else:
                    if Ktotal == 0:
                        self.depcalc(e, nr, i, t)
                    else:
                        self.B = (float(self.d["dikeent " + str(e) + nr].get()))/Ktotal
            if ce is None:
                self.need.append("You must set a Workload Distribution")
                self.depcalc(e, nr, i, t)
            else:
                self.depcalc(ce, cnr, ci, t)

    def calkerma(self, e, nr, i, t):
        if self.ent[f"entryd {e}{nr}{i}"].get() != "":
            D = float(self.ent[f"entryd {e}{nr}{i}"].get()) # Distance (entryd)
            if self.barn[f"lab_bar {e}{nr}"].cget("text") == "Floor": #from fig. 4.4 NCRP 147
                D += 1.7
            elif self.barn[f"lab_bar {e}{nr}"].cget("text") == "Ceiling":
                D += 0.5
            else:
                D += 0.3
        else:
            D = None
        # ======== Workload ===========
        if f"othwork {e}{nr}{i}" in self.d:
            if self.var[f"workv {e}{nr}{i}"].get() == 2: # Number of patient
                if self.d[f"numpapwe {e}{nr}{i}"].get() != '':
                    n = int(self.d[f"numpapwe {e}{nr}{i}"].get())  # Number of patients (numpapwe)
                else:
                    n =None
            elif self.var[f"workv {e}{nr}{i}"].get() == 1:# total Workload option
                ws = wb['Workload']
                for x in range(1, 13):
                    if self.var[f"vsexroom {e}{nr}{i}"].get() == ws['A' + str(x)].value:  # X-ray room selection (vsexroom)
                        if self.d[f"worentry {e}{nr}{i}"].get() != '':
                            n = float(self.d[f"worentry {e}{nr}{i}"].get()) / ws['B' + str(x)].value  # Workload entry (worentry)
                        else:
                            n = None
        else:
            n = None
        # ======== Occupancy Factor T ===========
        if self.var[f"vraoccup {e}{nr}"].get() == 1:  # Occupancy option (vraoccup)
            if self.d[f"occupentry {e}{nr}"].get() != '':
                T = float(self.d[f"occupentry {e}{nr}"].get())  # Occupancy entry (occupentry)
            else:
                T = None
        elif self.var[f"vraoccup {e}{nr}"].get() == 2:
            ws = wb['Occupancy Factor ( T )']
            for x in range(2, 32):
                if self.var[f"vselocation {e}{nr}"].get() == ws['A' + str(x)].value:  # Location selection (vselocation)
                    T = float(ws['B' + str(x)].value)
        else:
            T = None
        # ======== K1 "air kerma" ===========
        if self.var[f"radiob_w {e}{nr}{i}"].get() == 1:
        # primary unshielded air kerma
            if self.ent[f"entk {e}{nr}{i}"].get() != "":
                K1 = float(self.ent[f"entk {e}{nr}{i}"].get())  # Entered kerma value (entk)
            else:
                K1 = None
        elif self.var[f"radiob_w {e}{nr}{i}"].get() == 2:
            if self.var[f"unairkerv {e}{nr}{i}"].get() == 1: #Selects fron NCRP 147
                # Secondary air kerma calculations based on the 'airkerv' value
                ws = wb["Uns Air Kerma"]
                room_type = self.var[f"vsexroom {e}{nr}{i}"].get()  # Fetching X-ray room type
                if self.var[f"airkerv {e}{nr}{i}"].get() == 1:
                    # Leakage air kerma calculation
                    if self.var[f"radiob_leak {e}{nr}{i}"].get() == 1:
                        for j in range(1, 11):
                            if room_type == ws['A' + str(j)].value:
                                K1 = float(ws['G' + str(j)].value)  # Leak & side scatter
                    elif self.var[f"radiob_leak {e}{nr}{i}"].get() == 2:
                        for j in range(1, 11):
                            if room_type == ws['A' + str(j)].value:
                                K1 = float(ws['I' + str(j)].value)  # Leakage and Forward/ Backscatter (Ksec)
                    else:
                        for j in range(1, 11):
                            if room_type == ws['A' + str(j)].value:
                                K1 = float(ws['E' + str(j)].value)  # Only Leakage

                elif self.var[f"airkerv {e}{nr}{i}"].get() == 2:
                    # Side-Scatter
                    for j in range(1, 11):
                        if room_type == ws['A' + str(j)].value:
                            K1 = float(ws['F' + str(j)].value)
                elif self.var[f"airkerv {e}{nr}{i}"].get() == 3:
                    # Forward/ Backscatter
                    for j in range(1, 11):
                        if room_type == ws['A' + str(j)].value:
                            K1 = float(ws['H' + str(j)].value)
            elif self.var[f"unairkerv {e}{nr}{i}"].get() == 2:
                # User-entered air kerma
                if self.ent[f"entk {e}{nr}{i}"].get() != "":
                    K1 = float(self.ent[f"entk {e}{nr}{i}"].get())  # Entered kerma value (entk)
                else:
                    K1 = None
            else:
                K1 = None
        else:
            K1 = None
        # ======== Use Factor ===========
        if self.var[f"radiob_w {e}{nr}{i}"].get() == 1:
            if self.ent[f"use_ent {e}{nr}{i}"].get()!= "":
                Us = float(self.ent[f"use_ent {e}{nr}{i}"].get())  # Use factor (use_ent)
            else:
                Us =None
        elif self.var[f"radiob_w {e}{nr}{i}"].get() == 2:
            Us = 1
        else:
            Us = None
        #====Missed variables========
        self.need = []
        if n is None:
            self.need.append("Workload")
            print('w')
        if Us is None:
            self.need.append("Use Factor (U)")
            print('us')
        if T is None:
            self.need.append("Occupancy Factor (T)")
            print('T')
        if K1 is None:
            self.need.append("Unshielded Air Kerma")
            print('k')
        if D is None:
            self.need.append("Distance from the Source to the Barrier (m)")
            print("d")
        if n and Us and T and K1 and D is not None:
            # ========== Calculate Kerma ==========
            K = n * Us * T * K1 / (D ** 2)
            return K

        return None  # Return None if there are missing variables

    def depcalc(self, e, nr, i, t):  # Barrier calculations
        title_key = f"titleresul {e}{nr}"
        # Destroy existing material labels if they exist
        if self.d.get(title_key) is not None:
            for o in range(1, 7):
                mat_key = f"resmat {o}{e}"
                if self.res.get(mat_key) is not None:
                    self.res[mat_key].destroy()
        else:
            for o in range(1, 7):
                self.res[f"resmat {o}{e}"] = None

        # Iterate over materials and perform calculations
        for o in range(1, self.var[f"vnumbmat {e}{nr}"].get() + 1):
            material = self.var[f"vmater {e}{o}{nr}"].get()
            if self.need:
                if material == "Select Material":
                    self.need.append("Select Material")
                self.need = list(dict.fromkeys(self.need))
                self.thm[f"xbar {e}{o}{nr}"] = "\n".join(self.need)
                self.display_results(e, o, nr, t)
            else:
                if self.var[f"radiob_w {e}{nr}{i}"].get() == 1:
                    ws = wb['prim abc']
                    for x in range(2, 39):
                        if self.var[f"vsexroom {e}{nr}{i}"].get() == ws[f'A{x}'].value:
                            self.al = ws[f'B{x}'].value
                            self.bl = ws[f'C{x}'].value
                            self.cl = float(ws[f'D{x}'].value)
                            self.ac = ws[f'E{x}'].value
                            self.bc = ws[f'F{x}'].value
                            self.cc = float(ws[f'G{x}'].value)
                            self.ag = ws[f'H{x}'].value
                            self.bg = ws[f'I{x}'].value
                            self.cg = float(ws[f'J{x}'].value)
                            self.ast = ws[f'K{x}'].value
                            self.bs = ws[f'L{x}'].value
                            self.cs = float(ws[f'M{x}'].value)
                            self.apg = ws[f'N{x}'].value
                            self.bpg = ws[f'O{x}'].value
                            self.cpg = float(ws[f'P{x}'].value)
                            self.aw = ws[f'Q{x}'].value
                            self.bw = ws[f'R{x}'].value
                            self.cw = float(ws[f'S{x}'].value)

                    # Preshielding calculations
                    if self.var[f"preshvar {e}{nr}{i}"].get() == 1:
                        ws = wb["Equiv. thickness of prim pres"]
                        if self.var[f"radiob_pre {e}{nr}{i}"].get() == 1:
                            self.xlead = float(ws['B3'].value)
                            self.xconc = float(ws['C3'].value)
                            self.xsteel = float(ws['D3'].value)
                        elif self.var[f"radiob_pre {e}{nr}{i}"].get() == 2:
                            self.xlead = float(ws['B4'].value)
                            self.xconc = float(ws['C4'].value)
                            self.xsteel = float(ws['D4'].value)
                    else:
                        self.xlead, self.xconc, self.xsteel = 0, 0, 0

                elif self.var[f"radiob_w {e}{nr}{i}"].get() == 2:
                    self.xlead, self.xconc, self.xsteel = 0, 0, 0
                    ws = wb['sec abc']
                    for x in range(3, 18):
                        if self.var[f"vsexroom {e}{nr}{i}"].get() == ws[f'A{x}'].value:
                            self.al = ws[f'B{x}'].value
                            self.bl = ws[f'C{x}'].value
                            self.cl = float(ws[f'D{x}'].value)
                            self.ac = ws[f'E{x}'].value
                            self.bc = ws[f'F{x}'].value
                            self.cc = float(ws[f'G{x}'].value)
                            self.ag = ws[f'H{x}'].value
                            self.bg = ws[f'I{x}'].value
                            self.cg = float(ws[f'J{x}'].value)
                            self.ast = ws[f'K{x}'].value
                            self.bs = ws[f'L{x}'].value
                            self.cs = float(ws[f'M{x}'].value)
                            self.apg = ws[f'N{x}'].value
                            self.bpg = ws[f'O{x}'].value
                            self.cpg = float(ws[f'P{x}'].value)
                            self.aw = ws[f'Q{x}'].value
                            self.bw = ws[f'R{x}'].value
                            self.cw = float(ws[f'S{x}'].value)

                # Material thickness calculations
                if material == "Select Material":
                    self.thm[f"xbar {e}{o}{nr}"] = "Select Material"
                elif material == "Lead":
                    self.thm[f"xbar {e}{o}{nr}"] = (1 / (self.al * self.cl)) * math.log(
                        (self.B ** (-self.cl) + self.bl / self.al) / (1 + self.bl / self.al)) - self.xlead
                elif material == "Concrete":
                    self.thm[f"xbar {e}{o}{nr}"] = (1 / (self.ac * self.cc)) * math.log(
                        (self.B ** (-self.cc) + self.bc / self.ac) / (1 + self.bc / self.ac)) - self.xconc
                elif material == "Gypsum Wallboard":
                    self.thm[f"xbar {e}{o}{nr}"] = (1 / (self.ag * self.cg)) * math.log(
                        (self.B ** (-self.cg) + self.bg / self.ag) / (1 + self.bg / self.ag))
                elif material == "Steel":
                    self.thm[f"xbar {e}{o}{nr}"] = (1 / (self.ast * self.cs)) * math.log(
                        (self.B ** (-self.cs) + self.bs / self.ast) / (1 + self.bs / self.ast)) - self.xsteel
                elif material == "Plate Glass":
                    self.thm[f"xbar {e}{o}{nr}"] = (1 / (self.apg * self.cpg)) * math.log(
                        (self.B ** (-self.cpg) + self.bpg / self.apg) / (1 + self.bpg / self.apg))
                elif material == "Wood":
                    self.thm[f"xbar {e}{o}{nr}"] = (1 / (self.aw * self.cw)) * math.log(
                        (self.B ** (-self.cw) + self.bw / self.aw) / (1 + self.bw / self.aw))

                self.display_results(e, o, nr, t)

        # If there's no existing title result, create a new title
        if self.d.get(title_key) is None:
            self.op = 0
            self.d[title_key] = ttk.Label(self.d[f"resultframe {t}{nr}"], style="AL.TLabel",
                text=self.barn[f"lab_bar {e}{nr}"].cget("text") + ": ")
            self.d[title_key].grid(row=str(self.ep - self.op), column=0, pady=3, padx=3, sticky="s")
            self.d[f"spot {e}{nr}"] = self.ep
        else:
            # If it exists, update it
            self.d[title_key].destroy()
            self.d[title_key] = ttk.Label(self.d[f"resultframe {t}{nr}"], style="AL.TLabel",
                text=self.barn[f"lab_bar {e}{nr}"].cget("text") + ": ")
            self.d[title_key].grid(row=str(self.d[f"spot {e}{nr}"]), column=0, pady=3, padx=3, sticky="s")
        self.ep += 1

        for o in range(1, self.var[f"vnumbmat {e}{nr}"].get() + 1):
            if not hasattr(self, 'rdata'):
                self.rdata = {}
            # Store the results for further use
            if isinstance(self.thm[f'xbar {e}{o}{nr}'], float):
                barrier = self.barn[f"lab_bar {e}{nr}"].cget("text")
                material = self.var[f"vmater {e}{o}{nr}"].get()
                thickness = str(round(self.thm[f"xbar {e}{o}{nr}"], 3))
                # Ensure room number exists in rdata
                if f'{nr}' not in self.rdata:
                    self.rdata[f'{nr}'] = {}
                # Ensure barrier exists for the room
                if barrier not in self.rdata[f'{nr}']:
                    self.rdata[f'{nr}'][barrier] = {}
                # Add or update the material thickness without overwriting
                material_key = material + " (mm)"
                if material_key in self.rdata[f'{nr}'][barrier]:
                    # Material already exists, update its value
                    self.rdata[f'{nr}'][barrier][material_key] = thickness
                else:
                    # Add new material to the barrier
                    self.rdata[f'{nr}'][barrier][material_key] = thickness
            print(self.rdata)

    def depCTcal(self, e, nr, t):
        title_key = f"titleresul {e}{nr}"
        # Assign variable names for better readability
        dist_var = self.var[f"dist_var {e}{nr}"].get()
        bp_var = self.var[f"bp_var {t}"].get()
        hp_var = self.var[f"hp_var {t}"].get()
        sh_var = self.var[f"sh_var {e}{nr}"].get()
        dlpb_var = self.var[f"dlpb_var {t}"].get()
        dlph_var = self.var[f"dlph_var {t}"].get()
        numbodyp = self.var[f"numbodyscans {t}"].get() # the number of body phases
        numheadp = self.var[f"numheadscans {t}"].get() # the number of head phases
        kvp_var = self.var[f"kvp_var {t}"].get()
        T = self.var[f"occup {e}{nr}"].get()
        # Destroy existing material labels if they exist
        if self.d.get(title_key) is not None:
            for o in range(1, 7):
                mat_key = f"resmat {o}{e}"
                if self.res.get(mat_key) is not None:
                    self.res[mat_key].destroy()
        else:
            for o in range(1, 7):
                self.res[f"resmat {o}{e}"] = None

        for o in range(1, self.var[f"vnumbmat {e}{nr}"].get() + 1):
            material = self.var[f"vmater {e}{o}{nr}"].get()

            # Calculation condition check
            if (dist_var != "" and float(dist_var) != 0 and float(bp_var) != 0 and float(hp_var) != 0 and float(
                sh_var) != 0 and (float(dlpb_var) != 0 or float(dlph_var) != 0)):
                # Calculate total contributions for body scans
                total_scansb = 0
                k1_body = 1.2 * (3e-4) * float(dlpb_var)
                print(f'kbody:{k1_body}')
                for i in range(1, numbodyp + 1):
                    # Fetch the user-defined percentage for phase i
                    per_phase_body = self.var[
                        f"perbodyscans {t}{i - 1}"].get()  # Use i-1 because phases are zero-indexed the code
                    # Body scans contribution for phase i
                    total_scansb += (per_phase_body / 100) * float(bp_var)
                # Calculate total contributions for head scans
                total_scansh = 0
                k1_head = (9e-5) * float(dlph_var)
                print(f'khead:{k1_head}')
                for i in range(1, numheadp + 1):
                    # Fetch the user-defined percentage for phase i
                    per_phase_head = self.var[
                        f"perheadscans {t}{i - 1}"].get()
                    # Head scans contribution for phase i
                    total_scansh += (per_phase_head / 100) * float(hp_var)

                K_l = f"K {self.barn[f'lab_bar {e}{nr}'].cget('text')}{nr}"
                self.d[K_l] = (1 / (float(dist_var)** 2)) * float(T) *((k1_body* total_scansb) + (k1_head* total_scansh))
                print(f'Ksec:{self.d[K_l]}')
                B = float(sh_var) / float(self.d[K_l])
                print(f'B: {B}')

                # Thickness calculation
                if material == "Lead":
                    if kvp_var == 120:
                        a, b, c = 2.246, 5.73, 0.547
                    else:
                        a, b, c = 2.009, 3.99, 0.342
                elif material == "Concrete":
                    if kvp_var == 120:
                        a, b, c = 0.0383, 0.0142, 0.658
                    else:
                        a, b, c = 0.0336, 0.0122, 0.519
                else:
                    self.thm[f"xbar {e}{o}{nr}"] = "Select Material"
                    self.display_results(e, o, nr, t)
                    continue  # Skip to the next iteration if material is not selected

                self.thm[f"xbar {e}{o}{nr}"] = float((1 / (a * c)) * math.log((B ** -c + (b / a)) / (1 + (b / a))))
                self.display_results(e, o, nr, t)

            else:
                # Populate self.need based on missing or zero values
                self.need = []
                if dlpb_var == 0 and dlph_var == 0:
                    self.need.append("zero entries")
                
                if dist_var == "":
                    self.need.append("Distance")
                if bp_var == 0:
                    self.need.append("zero entries")
                if hp_var == 0:
                    self.need.append("zero entries")
                if sh_var == 0:
                    self.need.append("zero entries")
                if material == "Select Material":
                    self.need.append("Select Material")
                self.need = list(dict.fromkeys(self.need))
                self.thm[f"xbar {e}{o}{nr}"] = "\n".join(self.need)
                self.display_results(e, o, nr, t)

        # Title result handling
        if self.d.get(title_key) is None:
            self.op = 0
            self.d[title_key] = ttk.Label(self.d[f"resultframe {t}{nr}"], style="AL.TLabel",
                                          text=self.barn[f"lab_bar {e}{nr}"].cget("text") + ": ")
            self.d[title_key].grid(row=str(self.ep - self.op), column=0, pady=3, padx=3, sticky="s")
            self.d[f"spot {e}{nr}"] = self.ep
        else:
            self.d[title_key].destroy()
            self.d[title_key] = ttk.Label(self.d[f"resultframe {t}{nr}"], style="AL.TLabel",
                                          text=self.barn[f"lab_bar {e}{nr}"].cget("text") + ": ")
            self.d[title_key].grid(row=str(self.d[f"spot {e}{nr}"]), column=0, pady=3, padx=3, sticky="s")

        # Store the results for further use
        if isinstance(self.thm[f'xbar {e}{o}{nr}'], float):
            self.rdata = {
                self.barn[f"lab_bar {e}{nr}"].cget("text"): {material: str(round(self.thm[f"xbar {e}{o}{nr}"], 3))}}

        self.ep += 1

    def display_results(self, e, o, nr, t):
        title_key = f"titleresul {e}{nr}"
        spot_key = f"spot {e}{nr}"
        if spot_key not in self.d:
            self.d[spot_key] = self.ep  # or another default value
        if title_key not in self.d:
            self.d[title_key] = None
        if self.d[title_key] is None:
            if isinstance(self.thm[f'xbar {e}{o}{nr}'], str):
                if self.need:
                    # Calculate the number of elements in the list or string
                    num_elements = len(self.need)
                    self.res[f"resmat {o}{e}"] = Text(self.d[f"resultframe {t}{nr}"], height=num_elements, width=40,
                                                      borderwidth=0, background="#f7faf9", font='Helvetica 12')
                    # Insert "You must enter:" in black
                    self.res[f"resmat {o}{e}"].insert(END, "Invalid:", "black")
                    # Insert the rest of the text in red
                    self.res[f"resmat {o}{e}"].insert(END, f"\n{self.thm[f'xbar {e}{o}{nr}']}", "red")
                    # Configure the tags for text styles
                    self.res[f"resmat {o}{e}"].tag_configure("black", foreground="black")
                    self.res[f"resmat {o}{e}"].tag_configure("red", foreground='#f71616')
                    self.res[f"resmat {o}{e}"].config(state=DISABLED)
                else:
                    self.res[f"resmat {o}{e}"] = ttk.Label(self.d[f"resultframe {t}{nr}"], style="RL.TLabel",
                                                           text=f"{self.thm[f'xbar {e}{o}{nr}']}")
            else:
                self.res[f"resmat {o}{e}"] = ttk.Label(self.d[f"resultframe {t}{nr}"], style="AL.TLabel",
                                                       text=f"{self.var[f'vmater {e}{o}{nr}'].get()}: {round(self.thm[f'xbar {e}{o}{nr}'], 2)} mm")
            # Handle grid placement and manage row/column positions with op
            if o < 3:
                self.res[f"resmat {o}{e}"].grid(row=self.ep, column=o, pady=3, padx=3, sticky="w")
                self.op = 0  # Reset op if material is within the first 2 columns
            elif 2 < o < 5:
                if o == 3:
                    self.ep += 1  # Move to the next row for column 3
                    self.op += 1
                self.res[f"resmat {o}{e}"].grid(row=self.ep, column=o - 2, pady=3, padx=3, sticky="w")
            else:
                if o == 5:
                    self.ep += 1  # Move to a new row for column 5 and beyond
                    self.op += 1
                self.res[f"resmat {o}{e}"].grid(row=self.ep, column=o - 4, pady=3, padx=3, sticky="w")
        else:
            # If the title already exists, update it
            if isinstance(self.thm[f'xbar {e}{o}{nr}'], str):
                if self.need:
                    # Calculate the number of elements in the list or string
                    num_elements = len(self.need)
                    self.res[f"resmat {o}{e}"] = Text(self.d[f"resultframe {t}{nr}"], height=num_elements+1, width=40,
                                                      borderwidth=0, background="#f7faf9", font='Helvetica 12')
                    # Insert "You must enter:" in black
                    self.res[f"resmat {o}{e}"].insert(END, "Invalid:", "black")
                    # Insert the rest of the text in red
                    self.res[f"resmat {o}{e}"].insert(END, f"\n{self.thm[f'xbar {e}{o}{nr}']}", "red")
                    # Configure the tags for text styles
                    self.res[f"resmat {o}{e}"].tag_configure("black", foreground="black")
                    self.res[f"resmat {o}{e}"].tag_configure("red", foreground='#f71616')
                    self.res[f"resmat {o}{e}"].config(state=DISABLED)
                else:
                    self.res[f"resmat {o}{e}"] = ttk.Label(self.d[f"resultframe {t}{nr}"], style="RL.TLabel",
                                                           text=f"{self.thm[f'xbar {e}{o}{nr}']}")
            else:
                self.res[f"resmat {o}{e}"] = ttk.Label(self.d[f"resultframe {t}{nr}"], style="AL.TLabel",
                                                       text=f"{self.var[f'vmater {e}{o}{nr}'].get()}: {round(self.thm[f'xbar {e}{o}{nr}'], 2)} mm")

            # Use stored row information (spot) for placing result labels
            if o < 3:
                self.res[f"resmat {o}{e}"].grid(row=self.d[spot_key], column=o, pady=3, padx=3, sticky="w")
            elif 2 < o < 5:
                self.res[f"resmat {o}{e}"].grid(row=self.d[spot_key] + 1, column=o - 2, pady=3, padx=3, sticky="w")
            else:
                self.res[f"resmat {o}{e}"].grid(row=self.d[spot_key] + 2, column=o - 4, pady=3, padx=3, sticky="w")
